"""Parse CSV/XLSX spreadsheets into a common {columns, rows} shape."""

import csv
from io import StringIO
from pathlib import Path

SPREADSHEET_CONTENT_TYPES = frozenset({
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})


def _is_xlsx(path: Path, content_type: str) -> bool:
    return (
        content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or path.suffix.lower() == ".xlsx"
    )


def parse_spreadsheet(path: Path, content_type: str) -> dict:
    """Return {"columns": [...], "rows": [[str, ...], ...], "total_rows": int, "sheet_name": str | None}."""
    if _is_xlsx(path, content_type):
        return _parse_xlsx(path)
    return _parse_csv(path)


def _parse_csv(path: Path) -> dict:
    text = path.read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_name": None}

    # Sniff delimiter from the first 8k
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(StringIO(text), dialect)
    all_rows = [ [str(cell) for cell in row] for row in reader ]
    if not all_rows:
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_name": None}

    columns = all_rows[0]
    data_rows = all_rows[1:]
    return {
        "columns": columns,
        "rows": data_rows,
        "total_rows": len(data_rows),
        "sheet_name": None,
    }


def _parse_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active or (wb.worksheets[0] if wb.worksheets else None)
    if ws is None:
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_name": None}
    sheet_name = ws.title

    all_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(["" if cell is None else str(cell) for cell in row])

    wb.close()

    if not all_rows:
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_name": sheet_name}

    columns = all_rows[0]
    data_rows = all_rows[1:]
    return {
        "columns": columns,
        "rows": data_rows,
        "total_rows": len(data_rows),
        "sheet_name": sheet_name,
    }


def format_for_prompt(parsed: dict, filename: str, max_rows: int = 200) -> str:
    """Render a parsed spreadsheet as a markdown table for the Claude prompt."""
    columns = parsed.get("columns", [])
    rows = parsed.get("rows", [])
    total = parsed.get("total_rows", len(rows))
    sheet = parsed.get("sheet_name")

    header = f"[Spreadsheet: {filename}"
    if sheet:
        header += f" — sheet '{sheet}'"
    header += f" — {total} row{'s' if total != 1 else ''}]"

    if not columns and not rows:
        return f"{header}\n(empty file)"

    lines: list[str] = [header, ""]

    # Markdown table
    if columns:
        lines.append("| " + " | ".join(columns) + " |")
        lines.append("| " + " | ".join("---" for _ in columns) + " |")
    shown = rows[:max_rows]
    for r in shown:
        # Escape pipes so markdown table doesn't break
        safe = [str(c).replace("|", "\\|").replace("\n", " ") for c in r]
        lines.append("| " + " | ".join(safe) + " |")

    truncated = total - len(shown)
    if truncated > 0:
        lines.append("")
        lines.append(f"[... {truncated} more row{'s' if truncated != 1 else ''} not shown — user can view full table in the app]")

    return "\n".join(lines)
